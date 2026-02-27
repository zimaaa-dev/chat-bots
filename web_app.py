from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import yaml
import glob
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Конфигурация
CONFIG = {
    'FIRST_MESSAGE': 'Добро пожаловать! Для начала работы нажмите кнопку НАЧАТЬ',
    'FILENAME_MESSAGE': 'Введите название файла для сохранения результатов',
    'FILENAME_ID_MESSAGE': 'Ваши результаты сохранены под ID файла',
    'REQUEST_FILENAME': True,
    'REQUEST_USER': False,
    'TYPE_BUTTONS': 'reply'
}

FILES_START_DIRS = glob.glob(os.path.join('files_excel', 'files_start*'))

class QuestionState:
    def __init__(self):
        self.current_question = 0
        self.answers = {}
        self.last_idx = []
        self.used_vars = []
        self.final = False

def get_price_names():
    """Получает список доступных прайс-листов"""
    price_names = []
    for file_path in FILES_START_DIRS:
        price_names.append(os.path.basename(file_path).replace('files_start_', '').replace('.xlsx', ''))
    return price_names

def load_excel_data(filename):
    """Загружает данные из Excel файла"""
    try:
        df = pd.read_excel(filename, sheet_name='Otvet')
        return df
    except Exception as e:
        print(f"Ошибка загрузки файла {filename}: {e}")
        return None

def write_excel(filename, df):
    """Сохраняет ответы в Excel файл"""
    try:
        wb = load_workbook(filename, data_only=False)
        ws = wb['Otvet']
        
        for idx in range(len(df)):
            row_num = idx + 2
            
            # Обработка числового ответа
            num_val = df.iloc[idx]['Ответы число']
            if pd.isna(num_val):
                ws[f"C{row_num}"] = None
            else:
                try:
                    ws[f"C{row_num}"] = float(num_val)
                except (TypeError, ValueError):
                    ws[f"C{row_num}"] = num_val
            
            # Обработка текстового ответа
            text_val = df.iloc[idx]['Ответы текст']
            if pd.isna(text_val) or (isinstance(text_val, str) and text_val.strip() == ''):
                ws[f"D{row_num}"] = None
            else:
                ws[f"D{row_num}"] = str(text_val)
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Ошибка сохранения файла {filename}: {e}")
        return False

@app.route('/')
def index():
    """Главная страница"""
    return render_template('index.html', config=CONFIG, price_names=get_price_names())

@app.route('/start', methods=['POST'])
def start_questionnaire():
    """Начало опроса"""
    price_name = request.form.get('price_name')
    if not price_name:
        return jsonify({'error': 'Выберите прайс-лист'})
    
    # Загружаем данные
    filename = f"files_excel/files_start_{price_name}.xlsx"
    df = load_excel_data(filename)
    
    if df is None:
        return jsonify({'error': 'Файл не найден'})
    
    # Инициализируем состояние опроса
    session['question_state'] = QuestionState().__dict__
    session['current_file'] = filename
    session['df'] = df.to_dict()
    
    return redirect(url_for('question'))

@app.route('/question')
def question():
    """Текущий вопрос"""
    if 'question_state' not in session:
        return redirect(url_for('index'))
    
    state = QuestionState()
    state.__dict__.update(session['question_state'])
    df = pd.DataFrame(session['df'])
    
    if state.current_question >= len(df):
        return redirect(url_for('finish'))
    
    question_data = df.iloc[state.current_question]
    
    return render_template('question.html', 
                         question=question_data['Вопрос'],
                         question_num=state.current_question + 1,
                         total_questions=len(df),
                         keyboard_type=question_data['Клавиатура'],
                         options=question_data['Варианты ответов'],
                         data_type=question_data['Тип'])

@app.route('/answer', methods=['POST'])
def answer_question():
    """Обработка ответа"""
    if 'question_state' not in session:
        return redirect(url_for('index'))
    
    state = QuestionState()
    state.__dict__.update(session['question_state'])
    df = pd.DataFrame(session['df'])
    
    answer = request.form.get('answer')
    
    # Сохраняем ответ
    if answer:
        question_data = df.iloc[state.current_question]
        if question_data['Тип'] in ['int', 'float']:
            df.loc[state.current_question, 'Ответы число'] = float(answer)
        else:
            df.loc[state.current_question, 'Ответы текст'] = answer
    
    # Переходим к следующему вопросу
    state.current_question += 1
    session['question_state'] = state.__dict__
    session['df'] = df.to_dict()
    
    if state.current_question >= len(df):
        return redirect(url_for('finish'))
    
    return redirect(url_for('question'))

@app.route('/finish')
def finish():
    """Завершение опроса"""
    if 'question_state' not in session:
        return redirect(url_for('index'))
    
    state = QuestionState()
    state.__dict__.update(session['question_state'])
    df = pd.DataFrame(session['df'])
    
    # Сохраняем результаты
    filename = session['current_file']
    write_excel(filename, df)
    
    # Очищаем сессию
    session.clear()
    
    return render_template('finish.html', config=CONFIG)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)