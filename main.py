from dotenv import load_dotenv
import os
from aiogram import F, Dispatcher, Bot, Router
from aiogram.filters import Command
from aiogram import types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, FSInputFile, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import asyncio
import shutil
import pandas as pd
import numpy as np
import string
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
import subprocess
from PyPDF2 import PdfReader, PdfWriter
import yaml
from typing import List
import glob
import sys

PATH: os.PathLike = None
QUESTION_ID: int = None
FILENAME: os.PathLike = None
CONFIG = None
KEYBOARD_BUILDER = None
FILES_START_DIRS = glob.glob(os.path.join('files_excel', 'files_start*'))
PRICE_NAMES = None

load_dotenv()
token = os.getenv("TOKEN")
admin_id = os.getenv('ADMIN_ID')
pdf_id = os.getenv('PDF_ID')
bot = Bot(token=token)
dp = Dispatcher()
router = Router()
dp.include_router(router)


class Question(StatesGroup):
    current_question = State()

class FilenameState(StatesGroup):
    filename = State()

class MyKeyboard():
    def __init__(self, type_keyboard):
        self.type_keyboard = type_keyboard
    
    def create_button(self, value, callback = str | None):
        if self.type_keyboard.lower() == "inline":
            return InlineKeyboardButton(text=value, callback_data=callback)
        elif self.type_keyboard.lower() == 'reply':
            return KeyboardButton(text=value)
        else:
            raise ValueError('Не такого типа кнопки')

    def create_keyboard(self, buttons: List[List[KeyboardButton | InlineKeyboardButton]]):
        if self.type_keyboard.lower() == 'inline':
            return InlineKeyboardMarkup(
                inline_keyboard=buttons
            )
        elif self.type_keyboard.lower() == 'reply':
            return ReplyKeyboardMarkup(
                keyboard=buttons,
                resize_keyboard=True,
                one_time_keyboard=True
            )
        else:
            raise ValueError('Не такого типа кнопки')
        
async def run_soffice(*args):
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(
        None,
        lambda: subprocess.run(args, check=True)
    )


def write_excel(filename, df):
    wb = load_workbook(filename, data_only=False)
    ws = wb['Otvet']
    
    for idx in range(len(df)):
        row_num = idx + 2
        
        # Обработка числового ответа
        num_val = df.iloc[idx]['Ответы число']
        if pd.isna(num_val):
            ws[f"C{row_num}"] = None
        else:
            # Убедимся, что это число (а не строка)
            try:
                ws[f"C{row_num}"] = float(num_val)
            except (TypeError, ValueError):
                ws[f"C{row_num}"] = None

        # Обработка текстового ответа
        text_val = df.iloc[idx]['Ответы текст']
        if pd.isna(text_val) or (isinstance(text_val, str) and text_val.strip() == ''):
            ws[f"D{row_num}"] = None
        else:
            ws[f"D{row_num}"] = str(text_val)
            
    wb.save(filename)

def hidden_elements(ws, ws_formules):
    for idx, row in enumerate(ws.iter_rows(max_row=ws.max_row)):        
        if ws[f"A{idx + 1}"].value == 777:
            ws_formules.row_dimensions[idx + 1].hidden = True

    for idx, col in enumerate(ws.iter_cols(max_col=ws.max_column)):
        cell = ws.cell(1, idx + 1)
        if cell.value == 777:
            ws_formules.column_dimensions[get_column_letter(idx + 1)].hidden = True

@dp.message(Command("start"))
async def start(message):
    global CONFIG
    global KEYBOARD_BUILDER

    with open('config.yaml', 'r', encoding='UTF-8') as file:
        CONFIG = yaml.safe_load(file)

    KEYBOARD_BUILDER = MyKeyboard(type_keyboard=CONFIG['TYPE_BUTTONS'])
    keyboard = KEYBOARD_BUILDER.create_keyboard(
        [
            [KEYBOARD_BUILDER.create_button(value='НАЧАТЬ', callback='НАЧАТЬ')]
        ]
    )

    await message.answer(CONFIG['FIRST_MESSAGE'], reply_markup=keyboard)

@dp.message(Command("video"))
async def send_video_instruction(message: types.Message):
    """Обработчик команды /video — отправляет превью видео."""
    video_url = "https://rutube.ru/" 

    # Отправляем пояснение (опционально)
    await message.answer("🎥 Вот видеоинструкция по использованию бота:")

    # Отправляем ЧИСТУЮ ссылку — Telegram сам создаст превью
    await message.answer(video_url)

def create_price_keyboard():
    global PRICE_NAMES

    # Загружаем данные из YAML
    with open(os.path.join('files_excel', 'prices.yaml'), 'r', encoding='UTF-8') as f:
        data = yaml.safe_load(f)

    buttons = []

    for dir in FILES_START_DIRS:
        # Правильно получаем имя папки без путей (кроссплатформенно)
        dirname = os.path.basename(dir)
        
        # Проверяем, есть ли такой ключ в data
        if dirname not in data:
            raise KeyError(f"Не найден ключ '{dirname}' в prices.yaml")

        buttons.append([KeyboardButton(text=data[dirname])])
    
    buttons = buttons[::-1]

    # Сохраняем имена прайсов для дальнейшего использования
    PRICE_NAMES = list(data.values())

    return ReplyKeyboardMarkup(
        keyboard=buttons,
        resize_keyboard=True,
        one_time_keyboard=True
    )
    
@dp.message(lambda message: message.text == 'НАЧАТЬ')
async def start_button(update):
    global PATH
    global CONFIG

    if (isinstance(update, types.Message) and update.text == 'НАЧАТЬ') or (isinstance(update, types.CallbackQuery) and update.data == 'НАЧАТЬ'):
        user = update.from_user
        if isinstance(update, types.CallbackQuery):
            update = update.message

        with open('config.yaml', 'r', encoding='UTF-8') as file:
            CONFIG = yaml.safe_load(file)

        try:
            with open('access.yaml', 'r') as file:
                users = yaml.safe_load(file)
                if users is not None:
                    users_id = [u.get('user_id') for u in users]
                else:
                    users_id = []
        except:
            users_id = []
            print("Файл не найден!")

        username = str(user.id)
        if username not in users_id and "free" not in users_id:
            keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text='Отправить запрос на добавление', callback_data=f"ID:{username}")
                    ]
                ]
            )
            await update.answer(text='Доступ к данному боту закрыт', reply_markup=ReplyKeyboardRemove())
            await update.answer(text="Добавить?", reply_markup=keyboard)
        else:
            await update.answer(text='Загрука прайсов...', reply_markup=ReplyKeyboardRemove())
            dirs = os.listdir('files_excel')
            #soffice_path = r"D:\Progrmas\program\soffice.exe"

            if username not in dirs:
                for dir in FILES_START_DIRS:
                    # правильно получаем имя папки
                    dir_name = os.path.basename(dir)

                    user_dir = os.path.join('files_excel', username, dir_name)
                    os.makedirs(user_dir, exist_ok=True)

                    src_csv = os.path.join('files_excel', dir_name, 'CSV.csv')
                    dst_csv = os.path.join(user_dir, 'CSV.csv')

                    if not os.path.exists(src_csv):
                        raise FileNotFoundError(f"Файл не найден: {src_csv}")

                    shutil.copyfile(src_csv, dst_csv)

                    src_main = os.path.join('files_excel', dir_name, 'MAIN.xlsx')

                    subprocess.run(
                        [
                            #'xvfb-run',
                            'soffice',
                            '--headless',
                            '--convert-to', 'xlsx',
                            '--outdir', user_dir,
                            src_main
                        ],
                        check=True
                    )
                    shutil.copytree(os.path.join(dir, 'lists'), os.path.join(user_dir, 'lists'))
                    shutil.copytree(os.path.join(dir, 'images'), os.path.join(user_dir, 'images'))
                    

            # определяем рабочую папку пользователя
            dirs_user = os.listdir(os.path.join('files_excel', username))
            if "files_prem" in dirs_user:
                PATH = os.path.join('files_excel', username, 'files_prem')
            else:
                PATH = os.path.join('files_excel', username, 'files_start')
          
            await update.answer(
                'Выберите версию прайса',
                reply_markup=create_price_keyboard()
            )


@dp.message(lambda message: message.text in PRICE_NAMES)
async def change_price(update):
    global PATH

    with open(os.path.join("files_excel", 'prices.yaml'), 'r', encoding='UTF-8') as f:
        data = yaml.safe_load(f)
    
    for key, value in data.items():
        if value == update.text:
            dirname = os.path.basename(PATH)
            PATH = PATH.replace(dirname, key)
            #print(PATH)
            break
    else:
        await update.answer('Выберите версию прайса', reply_markup=create_price_keyboard())

    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text='ПРОДОЛЖИТЬ')],
            [KeyboardButton(text='Выбрать другой прайс')]            
        ],
        resize_keyboard=True,        # уменьшает кнопки под текст
        one_time_keyboard=True       # скрывает клавиатуру после нажатия
    )
    await update.answer(text=f"Вы выбрали - {update.text}", reply_markup=keyboard)

@dp.message(lambda message: message.text == 'Выбрать другой прайс' or message.text == 'ПРОДОЛЖИТЬ')
async def price(message, state):
    if message.text == 'Выбрать другой прайс':
        await message.answer('Выберите версию прайса', reply_markup=create_price_keyboard())
    else:
        await controller_questions(message, state)

@dp.callback_query(F.data.startswith("ID"))
async def send_id(query):
    try:
        with open('request.yaml', 'r') as file:
            users = yaml.safe_load(file)
            if users is not None:
                users_id = [u.get('user_id') for u in users]
            else:
                users_id = []
                users = []
    except:
        users_id = []
        print("Файл не найден!")
    
    id = query.data.split(":")[-1]
    if users is not None and id in users_id:
        await query.message.reply(text='Запрос уже был отправлен, ждите добавления')
    else:
        chat_id = query.message.chat.id
        user_chat = await bot.get_chat_member(chat_id=chat_id, user_id=id)
        username = user_chat.user.username
        full_name = user_chat.user.full_name
        await bot.send_message(chat_id=admin_id, text=f"""
        Новый пользователь просит добавления
<b>ID</b> - {id}
<b>username</b> {username}
<b>full name</b> {full_name}
t.me/{username}
        """
        , parse_mode=ParseMode.HTML)

        await query.message.reply(text='Запрос отправлен')
        users.append({'user_id': id, 'username': username})
        with open('request.yaml', 'w') as file:
            yaml.safe_dump(users, file)

    await bot.edit_message_reply_markup(
        chat_id=query.from_user.id,
        message_id=query.message.message_id,
        reply_markup=None
    )
    
async def controller_questions(message, state):
    global PATH

    main_xlsx = os.path.join(PATH, 'MAIN.xlsx')
    
    # Загружаем текущие данные
    answer_clean = pd.read_excel(main_xlsx, sheet_name='Otvet')
    
    # 🔥 Правильная очистка:
    answer_clean['Ответы число'] = np.nan      # числа → NaN
    answer_clean['Ответы текст'] = ''          # текст → пустая строка (или np.nan — без разницы)

    # Сохраняем очищенный файл
    write_excel(main_xlsx, answer_clean)

    # Загружаем заново (уже чистый)
    path = os.path.join(PATH, 'CSV.csv')
    questions = pd.read_csv(path, sep=';', index_col='id')
    answer_df = pd.read_excel(main_xlsx, sheet_name='Otvet')

    last_index = questions.index[-1]
    end_idx = last_index + 1

    await state.update_data(
        questions=questions,
        answer_df=answer_df,
        question_idx=0,
        last_idx=[],
        end_idx=end_idx,
        final=False
    )

    await send_question(message, state)
    await state.set_state(Question.current_question)



async def send_question(message, state):
    
    data = await state.get_data()
    df = data['questions']
    row = df.loc[data['question_idx']]
    vars = None

    data = await state.get_data()
    if data['final']:
        keyboard = create_random_keyboard()
        await message.answer(text='Вы хотите завершить?', reply_markup=keyboard)
    else:
        if row['Клавиатура'] == 1:
            buttons = [
                [
                    KEYBOARD_BUILDER.create_button('ДА', callback='ДА'),
                    KEYBOARD_BUILDER.create_button('НЕТ', callback='НЕТ'),
                ],
                [
                    KEYBOARD_BUILDER.create_button('НАЗАД', callback='НАЗАД'),
                    
                ],
                
            ]
            keyboard = KEYBOARD_BUILDER.create_keyboard(buttons)

        elif row['Клавиатура'] == 2:
            buttons = [
                [KEYBOARD_BUILDER.create_button('НАЗАД', callback='НАЗАД')],
                
            ]
            if row['Варианты ответов'] is not np.nan:
                vars = row['Варианты ответов'].split(',')
                vars = [var.strip() for var in vars]
                if len(vars) > 1:
                    pairs = []
                    for idx in range(0, len(vars), 2):
                        if idx + 1 == len(vars):
                            pairs.append([KEYBOARD_BUILDER.create_button(value=vars[idx], callback=vars[idx])
                            ])
                        else:
                            pairs.append(
                                [KEYBOARD_BUILDER.create_button(value=vars[idx],callback=vars[idx]),
                                 KEYBOARD_BUILDER.create_button(value=vars[idx + 1], callback=vars[idx + 1]),
                                 ]
                            )
                else:
                    pairs = [[KEYBOARD_BUILDER.create_button(value=var, callback=var)]for var in vars]
                buttons = pairs + buttons    

            keyboard = KEYBOARD_BUILDER.create_keyboard(buttons)

        elif row['Клавиатура'] == 0:
            buttons = []
            if row['Варианты ответов'] is not np.nan:
                vars = row['Варианты ответов'].split(',')
                vars = [var.strip() for var in vars]
                if len(vars) > 1:
                    pairs = []
                    for idx in range(0, len(vars), 2):
                        if idx + 1 == len(vars):
                            pairs.append([KEYBOARD_BUILDER.create_button(value=vars[idx], callback=vars[idx])
                            ])
                        else:
                            pairs.append(
                                [KEYBOARD_BUILDER.create_button(value=vars[idx],callback=vars[idx]),
                                 KEYBOARD_BUILDER.create_button(value=vars[idx + 1], callback=vars[idx + 1]),
                                 ]
                            )
                else:
                    pairs = [[KEYBOARD_BUILDER.create_button(value=var, callback=var)]for var in vars]
                buttons = pairs + buttons    

            keyboard = KEYBOARD_BUILDER.create_keyboard(buttons)

        elif row['Клавиатура'] == 5:
            # Новый вариант с "вычитанием" нажатых кнопок
            used_vars = data.get('used_vars', [])
            vars = []
            if row['Варианты ответов'] is not np.nan:
                vars = [var.strip() for var in row['Варианты ответов'].split(',')]
                # исключаем уже выбранные
                vars = [v for v in vars if v not in used_vars]

                # формируем кнопки в 2 столбика, как в вариантах 0 и 2
                if len(vars) > 1:
                    pairs = []
                    for idx in range(0, len(vars), 2):
                        if idx + 1 == len(vars):
                            pairs.append([KEYBOARD_BUILDER.create_button(value=vars[idx], callback=vars[idx])])
                        else:
                            pairs.append([
                                KEYBOARD_BUILDER.create_button(value=vars[idx], callback=vars[idx]),
                                KEYBOARD_BUILDER.create_button(value=vars[idx + 1], callback=vars[idx + 1]),
                            ])
                    var_buttons = pairs
                else:
                    var_buttons = [[KEYBOARD_BUILDER.create_button(value=var, callback=var)] for var in vars]
            else:
                var_buttons = []

            buttons = [
                [KEYBOARD_BUILDER.create_button('НЕТ', callback='НЕТ')],
            ] + var_buttons
         
            keyboard = KEYBOARD_BUILDER.create_keyboard(buttons)

            # сохраняем used_vars в state
            await state.update_data(used_vars=used_vars)


        # обновляем состояние вопроса
        await state.update_data(
            type_keyboard=row['Клавиатура'],
            type_data=row['Тип'],
            condition=row['Условие'],
            root=row['Переход'],
            input=row['Строка ввода'],
            vars=vars,
            start=row['Начало'],
            end=row['Конец'],
            answer=row['Ответ'],
            list_name=row['Список']
        )
        if not pd.isna(row['Картинка']):
            img_path = os.path.join(PATH, 'images', row['Картинка'])
            photo = FSInputFile(img_path)
            user_id = message.from_user.id
            await bot.send_photo(chat_id=user_id, photo=photo, caption=row['Описание'])
        await message.answer(text=row['Вопрос'], reply_markup=keyboard)
    
def create_random_keyboard():
    indices = [(0, 0), (0, 1), (1, 0), (1, 1), (2, 1), (2, 0)]
    np.random.shuffle(indices)
    buttons = [KEYBOARD_BUILDER.create_button(value='-', callback='НЕТ') for _ in range(6)]
    buttons[0] = KEYBOARD_BUILDER.create_button(value='ДА', callback='ДА')
    buttons[1] = KEYBOARD_BUILDER.create_button(value='НЕТ', callback='НЕТ')
    array = np.zeros(shape=(3, 2))
    array = array.tolist()
    for idx in range(6):
        indice = indices[-1]
        array[indice[0]][indice[1]] = buttons[idx]
        indices.pop()
    
    return KEYBOARD_BUILDER.create_keyboard(array)
    

async def exit(message, state):
    data = await state.get_data()
    answer = data['answer_df']

    write_excel(os.path.join(PATH, 'MAIN.xlsx'), answer)


    os.makedirs(os.path.join(PATH, 'temp'), exist_ok=True)
    #soffice_path = r"D:\Progrmas\program\soffice.exe"
    subprocess.run(['soffice', '--headless', '--invisible', '--convert-to', 'xlsx', '--outdir', os.path.join(PATH, 'temp'), os.path.join(PATH, 'MAIN.xlsx')], check=True)
    
    wb = load_workbook(os.path.join(PATH, 'temp', "MAIN.xlsx"), data_only=True)
    wb_formules = load_workbook(os.path.join(PATH, 'temp', "MAIN.xlsx"), data_only=False)
    for sn in wb.sheetnames:
        hidden_elements(wb[sn], wb_formules[sn])
    wb_formules.save(os.path.join(PATH, 'temp', "MAIN.xlsx"))


    await state.clear()
    if CONFIG['REQUEST_FILENAME']:
        await message.answer(CONFIG['FILENAME_MESSAGE'], reply_markup=ReplyKeyboardRemove())
        await state.set_state(FilenameState.filename)
    else:
        await message.answer(CONFIG['FILENAME_ID_MESSAGE'], reply_markup=ReplyKeyboardRemove())
        await convert2pdf(message)
                    

# @dp.message(Question.current_question)
async def answer(message, state: FSMContext):
    if isinstance(message, types.CallbackQuery):
        message_text = message.data
        message = message.message
    elif isinstance(message, types.Message):
        message_text = message.text
    else:
        message_text = None

    if message_text == "ЗАВЕРШИТЬ ПРОГРАММУ":
        await state.update_data(final=True)
        await send_question(message, state)
        return
        
    data = await state.get_data()
    question_id = data['question_idx']
    data = await state.get_data()
    answer = data['answer_df']
    df = data['questions']


        # --- обработка нового варианта клавиатуры 5 ---
    if data.get('type_keyboard') == 5:
        used_vars = data.get('used_vars', [])

        if message_text in (data.get('vars') or []):
            # запоминаем выбранный вариант
            used_vars.append(message_text)
            await state.update_data(used_vars=used_vars)

        if message_text in ['НЕТ', 'ДАЛЕЕ']:
            # сброс выбора
            await state.update_data(used_vars=[])


    if data['final']:
        if message_text == 'ДА':
            await exit(message, state)
            return
        elif message_text == 'НЕТ':
            await state.update_data(final=False)
            await send_question(message, state)
            return
        else:
            await message.answer("Выберите ДА или НЕТ!")
            await send_question(message, state)
            return
    if message_text == 'НАЗАД':
        question_id = data['last_idx'][-1]
        answer.loc[question_id, ['Ответы текст', 'Ответы число']] = np.nan
        last_idx = data['last_idx']
        if len(last_idx) == 1:
            last_idx = []
        else:
            last_idx.pop()
        await state.update_data(question_idx=question_id, last_idx=last_idx)
        await send_question(message, state)
        return
    if message_text == 'ДАЛЕЕ':
        end_idx = data['end_idx']  # Получаем end_idx из состояния
        if question_id + 1 < end_idx:
            last_idx = data['last_idx']
            last_idx.append(question_id)
            await state.update_data(question_idx=question_id + 1, last_idx=last_idx)
            await send_question(message, state)
            return
        else:
            await exit(message, state)
            return

    last_idx = data['last_idx']
    end_idx = data['end_idx']
    type_keyboard = data['type_keyboard']
    type_data = data['type_data']
    condition = data['condition']
    root = data['root']
    vars = data["vars"]
    input = data['input']
    start = data['start']
    end = data['end']
    answer_text = data['answer']
    list_name = data['list_name']

    if type_keyboard == 1 and message_text != "ДА" and message_text != "НЕТ":
        await message.answer("Выберите ДА или НЕТ!")
        await send_question(message, state)
        return 
    
    if type_keyboard == 2 and input == 'no':
        if message_text not in vars:
            await message.answer("Выберите из вариантов на клавиатуре")
            await send_question(message, state)
            return

    last_idx.append(question_id)

    if type_data == 'date':
        from datetime import datetime
        try:
            if not message_text == 'СЕГОДНЯ':
                datetime.strptime(message_text, "%d.%m.%Y")
            answer.loc[question_id, 'Ответы текст'] = message_text
        except ValueError:
            await message.answer("Введите дату в формате ДД.ММ.ГГГГ")
            await send_question(message, state)
            return 

    elif type_data == 'float':
        try:
            text = message_text
            text = text.split(',')
            if len(text) == 1:
                answer.loc[question_id, 'Ответы число'] = float(text[0])
            elif len(text) > 2:
                raise ValueError
            else:
                for id, num in enumerate(text):
                    if not num.strip().isdigit():
                        raise ValueError
                    text[id] = num.strip()
                text = ".".join(text)
                answer.loc[question_id, 'Ответы число'] = float(text)

        except ValueError:
            await message.answer("Некорректный формат ввода, введите число.")
            await send_question(message, state)
            return
        if not pd.isna(answer_text):
            if not (start <= answer.loc[question_id, 'Ответы число'] <= end):
                await message.answer(answer_text)
                await send_question(message, state)
                return 

    else:
        # Обработка str типа даннных (ввод разрешен)
        answer['Ответы текст'] = answer['Ответы текст'].astype(str)
        if not pd.isna(list_name):
            # Проверка наличия списка
            path2list = os.path.join(PATH, 'lists', list_name)
            names = pd.read_csv(path2list, header=None).iloc[:, 0].tolist()
            is_english = True
            for char in message_text:
                if char not in string.ascii_letters and not char.isdigit():
                    is_english = False
                    break
            if message_text.upper() in names and is_english:
                if message_text.isdigit():
                    ans = str(int(message_text))
                    answer.loc[question_id, 'Ответы текст'] = ans.upper()
                else:
                    answer.loc[question_id, 'Ответы текст'] = str(message_text).upper()
            else:
                await message.answer(answer_text)
                await send_question(message, state)
                return
        else:
            if message_text.isdigit():
                ans = str(int(message_text))
                answer.loc[question_id, 'Ответы текст'] = ans
            else:
                answer.loc[question_id, 'Ответы текст'] = str(message_text)

    
    next_question_id = question_id + 1

        # --- обработка переходов ---
    if root and not pd.isna(root):
        if not condition or pd.isna(condition):
            # Условие пустое, но переход задан → берём первый переход
            try:
                next_question_id = int(str(root).split(',')[0].strip())
            except Exception:
                pass
        else:
            # Условие заполнено → проверяем соответствие
            condition_list = [elem.strip() for elem in str(condition).split(',')]
            jump_to = [int(elem.strip()) for elem in str(root).split(',')]
            for cond, jump in zip(condition_list, jump_to):
                if message_text == cond:
                    next_question_id = jump
                    break

            
    if next_question_id >= end_idx:  # Изменено условие, чтобы проверять >= end_idx
        await state.clear()
        await message.answer("Вопросов больше нет!")
    else:
        await state.update_data(question_idx=next_question_id, last_idx=last_idx)
        await send_question(message, state)

@dp.message(FilenameState.filename)
async def get_filename(message, state):
    filename = message.text
    await state.clear()
    await convert2pdf(message, filename)

async def convert2pdf(message, filename=None):
    if filename is None:
        filename = str(message.from_user.id)

    input_file = os.path.abspath(os.path.join(PATH, 'temp', "MAIN.xlsx"))
    try:
        #soffice_path = r"D:\Progrmas\program\soffice.exe"
        subprocess.run(['soffice', '--headless', '--invisible', '--convert-to', 'pdf', '--outdir', PATH, input_file], check=True)
        reader = PdfReader(os.path.join(PATH, 'MAIN.pdf'))
        writer = PdfWriter()
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if str(CONFIG['SEPARATOR']) in text:
                break
            else:
                writer.add_page(page)
        with open(os.path.join(PATH, f'{filename}.pdf'), "wb") as f:
            writer.write(f)
        await message.answer(CONFIG['LAST_MESSAGE'])
        file = FSInputFile(os.path.join(PATH, f"{filename}.pdf"))
        global FILENAME
        FILENAME = os.path.join(PATH, f"{filename}.pdf")
        keyboard = None
        if CONFIG['SEND_FILE']:
            keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text='Отправить файл?', callback_data=f'send_file')
                    ],
                    [
                        InlineKeyboardButton(text='Начать заново', callback_data=f'start')
                    ]
                ],
                resize_keyboard=True
            )
        await message.answer_document(file, caption='Расчет PDF', reply_markup=keyboard)

        if not CONFIG['SEND_FILE']:
            os.remove(FILENAME)
    except subprocess.CalledProcessError as e:
        await message.answer(f"Ошибка конвертации {e}")

@dp.callback_query(F.data.startswith("send_file"))
async def check_keyborad(query):
    file = FSInputFile(FILENAME)
    user_id = pdf_id
    await bot.send_document(chat_id=user_id, document=file, caption='Ваш файл')
    os.remove(FILENAME)
    await start(query.message)

@dp.callback_query(F.data.startswith("start"))
async def check_start(query):
    os.remove(FILENAME)
    await start(query.message)

async def main():
    router.message(lambda message: message.text == 'НАЧАТЬ')(start_button)
    router.callback_query(lambda query: query.data == 'НАЧАТЬ')(start_button)
    router.message(Question.current_question)(answer)
    router.callback_query(Question.current_question)(answer)
    router.message(Command("video"))(send_video_instruction)
    await dp.start_polling(bot)

if __name__ == "__main__":
    import sys
    import asyncio

    # 👇 ЭТА СТРОЧКА — КЛЮЧЕВАЯ ДЛЯ WINDOWS
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    print("Бот запущен!")
    asyncio.run(main())